#!/bin/env python3

#DATABASE
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

#colorama
from colorama import Fore, Back, Style

#selenium
import selenium
from selenium import webdriver

class Main():
    def __init__(self):
        '''inititalize'''
        self.__author__ = "__th3end0f3v4ngl310n__"
        self.green = 'green'
        self.red = 'red'
        self.yellow = 'yellow'
        self.city = 'salavat'

        self.titles_array = []
        self.adress_array = []
        self.prices_array = []

        self.driver = webdriver.Chrome()
    def avito(self):
        '''parse'''

        def dl(text,color):
            '''debug log'''
            if color == 'yellow':
                print('~ ',Fore.YELLOW + '{}'.format(text), Style.RESET_ALL)
            elif color =='green':
                print('[',Fore.GREEN + 'OK', Style.RESET_ALL + '] ', text)
            elif color == 'red':
                 print('[',Fore.RED + 'ERROR',Style.RESET_ALL+ '] ', text)
            else:
                pass

        def create_db():
            '''create database'''
            wb = Workbook()
            sheet = wb.sheetnames
            wb.save('database.xlsx')
            dl('База данных сделана!', self.green)

        def parse():
            ''''parse from avito'''
            url = 'https://avito.ru/{}/sobaki'.format(self.city)

            dl('Селениум загрузился',self.green)
            self.driver.get(url)
            dl('Ссылка введена', self.green)

            titles = self.driver.find_elements_by_class_name('title-root-395AQ')
            for title in titles:
                self.titles_array.append(title.text)

            adress = self.driver.find_elements_by_class_name('geo-address-9QndR')
            for adres in adress:
                self.adress_array.append(adres.text)

            prices = self.driver.find_elements_by_class_name('price-text-1HrJ_')
            for price in prices:
                self.prices_array.append(price.text)

            print(self.adress_array)
            print(self.titles_array)
            print(self.prices_array)
            dl('Всё спаршено!', self.green)

        def close_drivers():
            for i in range(5):
                try:
                    self.driver.close()
                except:
                    break
                    print('COMLETED!')

        def fill_database():
            wb = load_workbook('database.xlsx')
            sheet = wb['Sheet']
            max_a = sheet.max_row
            max_b = sheet.max_row
            max_c = sheet.max_row
            max_d = sheet.max_row

            for i in self.titles_array:
                sheet.cell(row = max_a,column = 1).value = i
                max_a += 1

            for i in self.adress_array:
                sheet.cell(row = max_b, column = 2).value = i
                max_b += 1

            for i in self.prices_array:
                sheet.cell(row = max_c,column = 3).value = i
                max_c += 1

            wb.save('database.xlsx')
                #Call Functions
        create_db()
        parse()
        close_drivers()
        fill_database()

if __name__ == '__main__':
    root = Main()
    root.avito()
